import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import openpyxl.styles
import time
import os
import xlwings as xw
from xlwings.constants import HAlign

# ------------------------------------------------------------
# PROGRAMA 1: EXTRAI ALT, USP E SKU
# ------------------------------------------------------------
def iniciar_extracao():
    botao.config(text="Peraí...")
    botao.config(state="disabled")
    # Atualizar a interface gráfica para refletir as mudanças imediatamente
    root.update()

    # Variáveis para as URLs
    url_pagina = url_entry.get()
    login_url_aem = login_url_entry.get()
    login_url_qa = "https://p6-qa.samsung.com/aemapi/user/login_sso?s_user_email=felipe.sf@samsung.com"

    # Variáveis para os seletores
    seletor_alt = "#benefit img"
    seletor_usp = "#content > div > div > div.pd-g-product-detail-kv.aem-GridColumn.aem-GridColumn--default--12 > div.product-detail-kv > div.product-detail-kv__buying-tool > section > div.pd-buying-tool__info > div > div.pd-info__usp-list"
    seletor_sku = "pd-info__sku-code"

    # Verificar se os campos estão vazios
    if not url_pagina or not login_url_aem:
        messagebox.showwarning("Aviso", "Ambos os campos devem ser preenchidos.")
        botao.config(text="Bora!")
        botao.config(state="normal")
        return

    # Configurar opções para o modo headless
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    driver = webdriver.Edge(options=options)

    # Login QA
    try:
        driver.get(login_url_qa)
        time.sleep(1)
        current_url_qa = driver.current_url
        progress_bar["value"] = 20
        progress_bar.update()
    except:
        mostrar_mensagem("Erro", "Erro ao tentar logar no link de QA.")
        resetar_interface()
        return

    # Verificar se a URL atual não é válida
    if current_url_qa != "https://p6-qa.samsung.com/sites/":
        mostrar_mensagem("Erro", "Erro ao tentar logar no link de QA.")
        resetar_interface()
        return

    # Login AEM
    try:
        driver.get(login_url_aem)
        time.sleep(1)
        current_url_aem = driver.current_url
        progress_bar["value"] = 40
        progress_bar.update()
    except:
        mostrar_mensagem("Erro", "Erro ao tentar logar no AEM.")
        resetar_interface()
        return

    # Verificar se a URL atual não é válida
    if current_url_aem != "https://p6-us-author.samsung.com/sites.html/content" and current_url_aem != "https://p6-eu-author.samsung.com/sites.html/content" and current_url_aem != "https://p6-ap-author.samsung.com/sites.html/content":
        mostrar_mensagem("Erro", "Erro ao tentar logar no AEM.")
        resetar_interface()
        return

    # Acessar a URL da página
    try:
        driver.get(url_pagina)
        time.sleep(1)
        progress_bar["value"] = 60
        progress_bar.update()
    except:
        mostrar_mensagem("Erro", "Erro ao acessar a URL da página.")
        resetar_interface()
        return

    # Extrair dados
    try:
        # Extrair textos ALT
        alt_textos = [img.get_attribute("alt") for img in driver.find_elements(By.CSS_SELECTOR, seletor_alt)]
        alt_textos = list(set(alt_textos)) # Remover duplicatas

        # Tentar extrair SKU
        try:
            sku = driver.find_element(By.CLASS_NAME, seletor_sku).text

            # Extrair textos USP (tentar)
            try:
                usp_textos = driver.find_element(By.CSS_SELECTOR, seletor_usp).text.splitlines()
            except:
                usp_textos = []  # Lista vazia se não encontrar USP
        except:
            # Se não encontrar SKU, gerar nova URL
            nova_url = gerar_nova_url(url_pagina)
            driver.get(nova_url)
            time.sleep(1)

            # Extrair SKU e USP da nova URL
            try:
                usp_textos = driver.find_element(By.CSS_SELECTOR, seletor_usp).text.splitlines()
                sku = driver.find_element(By.CLASS_NAME, seletor_sku).text
            except:
                mostrar_mensagem("Erro", "Erro ao encontrar SKU ou USP na página.")
                resetar_interface()
                return
        
        progress_bar["value"] = 80
        progress_bar.update()
    except:
        mostrar_mensagem("Erro", "Erro ao extrair dados da página.")
        resetar_interface()
        return

    # Salvar dados na planilha
    wb = Workbook()
    progress_bar["value"] = 100
    progress_bar.update()
    wb.remove(wb.active)

    # Definir estilo de borda
    borda_fina = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                        bottom=openpyxl.styles.Side(style='thin'),
                                        left=openpyxl.styles.Side(style='thin'),
                                        right=openpyxl.styles.Side(style='thin'))

    if alt_textos:
        ws_alt = wb.create_sheet("ALT")
        # Configurar formatação para a aba "ALT"
        ws_alt.column_dimensions['A'].width = 100
        ws_alt.column_dimensions['B'].width = 100
        ws_alt.freeze_panes = 'A2'
        ws_alt['A1'].value = "English Text"
        ws_alt['A1'].font = ws_alt['B1'].font = openpyxl.styles.Font(bold=True)
        ws_alt['A1'].alignment = ws_alt['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws_alt['A1'].fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        ws_alt['A1'].border = borda_fina
        ws_alt['B1'].value = "Translated Text"
        ws_alt['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws_alt['B1'].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws_alt['B1'].border = borda_fina
                
        # Adicionar bordas apenas às células na Coluna A que possuem texto
        for row in range(1, len(alt_textos) + 1):
            texto = alt_textos[row - 1]
            ws_alt.cell(row=row + 1, column=1).value = texto
            if texto:
                ws_alt.cell(row=row + 1, column=1).border = borda_fina
                # Adicionar bordas em cada célula da Coluna B em que a célula da Coluna A tenha texto
                ws_alt.cell(row=row + 1, column=2).border = borda_fina

        # Definir formatação wrap e alinhamento vertical superior para todas as células a partir da linha 2
        for row in ws_alt.iter_rows(min_row=2, min_col=1, max_row=ws_alt.max_row, max_col=ws_alt.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')

    if usp_textos:
        ws_usp = wb.create_sheet("USP")
        # Configurar formatação para a aba "USP"
        ws_usp.column_dimensions['A'].width = 100
        ws_usp.column_dimensions['B'].width = 100
        ws_usp.freeze_panes = 'A2'
        ws_usp['A1'].value = "English Text"
        ws_usp['A1'].font = ws_usp['B1'].font = openpyxl.styles.Font(bold=True)
        ws_usp['A1'].alignment = ws_usp['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws_usp['A1'].fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        ws_usp['A1'].border = borda_fina
        ws_usp['B1'].value = "Translated Text"
        ws_usp['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws_usp['B1'].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws_usp['B1'].border = borda_fina

        # Adicionar textos USP na planilha
        for row, texto in enumerate(usp_textos, start=1):
            ws_usp.cell(row=row + 1, column=1).value = texto
            # Adicionar bordas apenas às células na Coluna A que possuem texto e às da Coluna B para cada uma da Coluna A que possua texto
            if texto:
                ws_usp.cell(row=row + 1, column=1).border = borda_fina
                ws_usp.cell(row=row + 1, column=2).border = borda_fina

        # Definir formatação wrap para todas as células a partir da linha 2
        for row in ws_usp.iter_rows(min_row=2, min_col=1, max_row=ws_usp.max_row, max_col=ws_usp.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')

    # Salvar arquivo
    nome_arquivo = f"{sku}_ALT-USP.xlsx"
    arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=nome_arquivo)
    if arquivo:
        salvar_arquivo(wb, arquivo)
    else:
        # mostrar_mensagem("Cancelado", "O salvamento do arquivo foi cancelado.")
        pass

    resetar_interface()

def gerar_nova_url(url):
    if url.startswith("https://www.samsung.com/") or url.startswith("https://p6-qa.samsung.com/"):
        return url + "/buy"
    elif url.startswith("https://p6-us-author.samsung.com/content/samsung/") or url.startswith("https://p6-eu-author.samsung.com/content/samsung/") or url.startswith("https://p6-ap-author.samsung.com/content/samsung/"):
        return url.replace(".html?wcmmode=disabled", "/buy.html?wcmmode=disabled")
    else:
        return None

def mostrar_mensagem(titulo, mensagem):
    messagebox.showinfo(titulo, mensagem)

def resetar_interface():
    botao.config(text="Bora!")
    botao.config(state="normal")
    progress_bar["value"] = 0

def salvar_arquivo(wb, nome_arquivo):
    try:
        wb.save(nome_arquivo)
        mostrar_mensagem("Sucesso", "Prontinho :)")
    except PermissionError:
        mostrar_mensagem("Erro", "O arquivo está aberto e não pode ser sobrescrito. Por favor, feche o arquivo e tente novamente.")

# ------------------------------------------------------------
# PROGRAMA 2: EDITA E GERA OS ARQUIVOS FINAIS
# ------------------------------------------------------------
def aplicar_formatacao_features(sheet):
    # Definir a cor de preenchimento amarela como uma tupla RGB
    fill_yellow = (255, 255, 0)  # Esta é a cor amarela em RGB
    # Encontrar a célula com o texto "Translated Text" e pintar de amarelo
    translated_text_row = None
    for row in range(1, sheet.api.UsedRange.Rows.Count + 1):
        if sheet.range((row, 5)).value == "Translated Text":
            sheet.range((row, 5)).color = fill_yellow
            translated_text_row = row
            break

    # Excluir o conteúdo da coluna E, exceto a célula "Translated Text"
    for row in range(2, sheet.api.UsedRange.Rows.Count + 1):
        if row != translated_text_row:
            sheet.range((row, 5)).value = "" 

    # Aplicar estilo negrito e centralizado à célula "Translated Text"
    if translated_text_row:
        cell = sheet.range((translated_text_row, 5))
        cell.api.Font.Bold = True
        cell.api.HorizontalAlignment = HAlign.xlHAlignCenter

def aplicar_formatacao_specs(sheet, filename):
    # Inserir uma nova coluna após a coluna K (11ª coluna)
    sheet.api.Columns(12).Insert()
    # Iterar sobre as linhas para aplicar borda na coluna L caso haja texto na célula da coluna M
    for row in range(1, sheet.api.UsedRange.Rows.Count + 1):
        # Verificar se a célula na coluna M tem texto
        if sheet.range((row, 13)).value:
            # Aplicar borda na célula correspondente na coluna L
            cell_L = sheet.range((row, 12))
            cell_L.api.Borders.LineStyle = 1  # xlContinuous
            cell_L.api.Borders.Weight = 2  # xlThin

    # Encontrar a célula com o texto "Value" e aplicar as modificações
    for row in range(1, sheet.api.UsedRange.Rows.Count + 1):
        if sheet.range((row, 11)).value == "Value":
            cell_L = sheet.range((row, 12))
            cell_L.value = "Translated Text"
            cell_L.color = (255, 255, 0)  # Definir a cor de preenchimento amarela
            cell_L.api.Font.Bold = True
            cell_L.api.HorizontalAlignment = HAlign.xlHAlignCenter
            cell_L.api.Borders.LineStyle = 1  # xlContinuous
            cell_L.api.Borders.Weight = 2  # xlThin

    # Renomear a aba com o prefixo do arquivo original
    prefix = filename.split("_SPECS")[0]  # Obter o prefixo do arquivo original
    sheet.name = prefix  # Renomear a aba com o prefixo

def editar_planilhas():
    # Abrir a caixa de diálogo para escolher os arquivos (permitir múltiplos)
    filepaths = filedialog.askopenfilenames(
        filetypes=[("Arquivos Excel", ("*.xlsx", "*.xls"))],
        multiple=True  # Habilitar a seleção múltipla
    )

    # Se nenhum arquivo foi selecionado, sair da função
    if not filepaths:
        return

    # Contadores para os diferentes tipos de arquivos
    specs_files = []
    features_file = None
    alt_usp_files = []
    for filepath in filepaths:
        filename = os.path.basename(filepath)
        if filename.endswith("_SPECS.xlsx") or filename.endswith("_SPECS.xls"):
            specs_files.append(filepath)
        elif filename.endswith("_FEATURES.xlsx") or filename.endswith("_FEATURES.xls"):
            if features_file:
                messagebox.showerror("Erro", "Selecione no máximo um arquivo '_FEATURES'.")
                return 
            features_file = filepath
        elif filename.endswith("_ALT-USP.xlsx") or filename.endswith("_ALT-USP.xls"):
            alt_usp_files.append(filepath) 

    # Verificar se há pelo menos um arquivo _SPECS
    if len(specs_files) == 0:
        messagebox.showerror("Erro", "Selecione pelo menos um arquivo '_SPECS'.")
        return 

    # Criar arquivo _Copy_Deck_Specs
    prefixo = os.path.basename(specs_files[0]).split("_")[0]
    copy_deck_specs_path = os.path.join(os.path.dirname(specs_files[0]), f"{prefixo}_Copy_Deck_Specs.xlsx")
    wb_copy_deck_specs = xw.Book()
    wb_copy_deck_specs.sheets[0].name = "Temp"

    # Criar arquivo _Copy_Deck_Features
    if features_file:
        prefixo = os.path.basename(features_file).split("_")[0]
        copy_deck_features_path = os.path.join(os.path.dirname(features_file), f"{prefixo}_Copy_Deck_Features.xlsx")
        wb_copy_deck_features = xw.Book()
        wb_copy_deck_features.sheets[0].name = "Temp"

        # Etapa 1: Criar cópias temporárias e aplicar formatação com macros xlwings
        for filepath in [features_file] + specs_files + alt_usp_files: 
            filename = os.path.basename(filepath) 
            # Criar cópia com sufixo "_TEMP" 
            temp_filepath = filepath[:-5] + "_TEMP.xlsx" 
            xw.Book(filepath).save(temp_filepath) 
            # Abrir a cópia e aplicar formatação 
            wb_temp = xw.Book(temp_filepath) 
            sheet = wb_temp.sheets.active 
            # Aplicar formatação dependendo do tipo de arquivo 
            if filename.endswith("_FEATURES.xlsx") or filename.endswith("_FEATURES.xls"):
                aplicar_formatacao_features(sheet) 
                for sheet in wb_temp.sheets:
                    sheet.api.Copy(After=wb_copy_deck_features.sheets[0].api) 
            elif filename.endswith("_SPECS.xlsx") or filename.endswith("_SPECS.xls"):
                aplicar_formatacao_specs(sheet, filename) 
                for sheet in wb_temp.sheets:
                    sheet.api.Copy(After=wb_copy_deck_specs.sheets[0].api)
            elif filename.endswith("_ALT-USP.xlsx") or filename.endswith("_ALT-USP.xls"):
                for sheet in wb_temp.sheets:
                    sheet.api.Copy(After=wb_copy_deck_features.sheets[0].api)
            wb_temp.close()
            os.remove(temp_filepath)  # Remover cópia temporária

        # Finalizar arquivo _Copy_Deck_Features 
        wb_copy_deck_features.sheets[0].delete()
        wb_copy_deck_features.save(copy_deck_features_path)

        # Reordenar as abas para garantir que a aba "Feature" seja a primeira
        for sheet in wb_copy_deck_features.sheets:
            if sheet.name == "Feature":
                sheet.api.Move(Before=wb_copy_deck_features.sheets[0].api)
        wb_copy_deck_features.save(copy_deck_features_path)

    # Finalizar arquivo _Copy_Deck_Specs
    wb_copy_deck_specs.sheets[0].delete()
    wb_copy_deck_specs.save(copy_deck_specs_path)

    messagebox.showinfo("Sucesso", "Arquivos finais gerados com sucesso!")

# ------------------------------------------------------------
# INTERFACE GRÁFICA COM TKINTER
# ------------------------------------------------------------
root = tk.Tk()
root.title("Copy Deck Preparer 3.0")
root.resizable(False, False)  # Impede redimensionamento da janela em largura e altura

# ------------------------------------------------------------
# Parte da interface gráfica do Programa 1
# ------------------------------------------------------------
frame_extrator_dados = tk.LabelFrame(root, text="ALT & USP Extractor", padx=10, pady=10)
frame_extrator_dados.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

# Campos de entrada
label_url = tk.Label(frame_extrator_dados, text="*URL da Página:")
label_url.grid(row=0, column=0, padx=5, pady=5)
url_entry = tk.Entry(frame_extrator_dados, width=50)
url_entry.grid(row=0, column=1, padx=5, pady=5)

label_url_login = tk.Label(frame_extrator_dados, text="*AEM Login URL:")
label_url_login.grid(row=1, column=0, padx=5, pady=5)
login_url_entry = tk.Entry(frame_extrator_dados, width=50)
login_url_entry.insert(0, "https://p6-us-author.samsung.com/aemapi/user/login_sso?s_user_email=felipe.sf@samsung.com") 
login_url_entry.grid(row=1, column=1, padx=5, pady=5)

# Botão de ação
botao = tk.Button(frame_extrator_dados, text="Bora!", command=iniciar_extracao)
botao.grid(row=2, columnspan=2, pady=10)

# Barra de progresso
progress_bar = ttk.Progressbar(frame_extrator_dados, orient="horizontal", length=200, mode="determinate")
progress_bar.grid(row=3, column=0, columnspan=2, pady=10)

# ------------------------------------------------------------
# Parte da interface gráfica do Programa 2
# ------------------------------------------------------------
frame_editor_planilhas = tk.LabelFrame(root, text="Copy Deck Editor", padx=10, pady=10)
frame_editor_planilhas.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

button_editar_planilhas = tk.Button(frame_editor_planilhas, text="Selecionar Planilhas", command=editar_planilhas)
button_editar_planilhas.pack(pady=60)

root.mainloop()
